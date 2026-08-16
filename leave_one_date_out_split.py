#!/usr/bin/env python3


from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_INPUT = Path(
    r"E:\YGWLAST\UAV-aphid data\2024ROI-sub-npy\label-2024.xlsx"
)

FILENAME_PATTERN = re.compile(
    r"^(?P<date>\d{8})-"          
    r"(?P<height>[^-]+)-"          
    r"(?P<plot_no>\d+)_"         
    r"(?P<repeat_no>\d+)_"       
    r"(?P<field_no>\d+)"          
    r"(?:\.[^.]+)?$"            
)

SPLIT_FILENAMES = {
    "train": "train_set_2024.xlsx",
    "val": "val_set_2024.xlsx",
    "test": "test_set_2024.xlsx",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="每个拍摄日期轮流作为一次测试集，生成留一日期法交叉验证数据。"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"输入 Excel，默认：{DEFAULT_INPUT}",
    )
    parser.add_argument(
        "--sheet",
        default="0",
        help="工作表名称或从 0 开始的序号，默认：0",
    )
    parser.add_argument(
        "--filename-col",
        default="filename",
        help="文件名列名，默认：filename",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="输出目录，默认在输入文件旁创建 leave_one_date_out_2024",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="允许覆盖脚本生成的同名文件",
    )
    return parser.parse_args()


def parse_sheet(value: str) -> str | int:
    value = value.strip()
    return int(value) if value.isdigit() else value


def extract_dates(df: pd.DataFrame, filename_col: str) -> pd.Series:
    """严格解析 filename，并返回 YYYYMMDD 字符串。"""
    if filename_col not in df.columns:
        raise ValueError(
            f"找不到 {filename_col!r} 列，当前列名为：{list(df.columns)!r}"
        )

    filenames = df[filename_col].astype("string").str.strip()
    extracted = filenames.str.extract(FILENAME_PATTERN)
    parsed_dates = pd.to_datetime(
        extracted["date"], format="%Y%m%d", errors="coerce"
    )
    invalid = filenames.isna() | extracted["date"].isna() | parsed_dates.isna()

    if invalid.any():
        examples = []
        for position in list(df.index[invalid][:10]):
            examples.append(
                f"Excel 第 {int(position) + 2} 行：{filenames.loc[position]!r}"
            )
        raise ValueError(
            "发现空 filename、非法日期或不符合格式的文件名：\n  "
            + "\n  ".join(examples)
        )

    if filenames.duplicated().any():
        duplicates = filenames[filenames.duplicated(keep=False)].unique()[:10]
        raise ValueError(f"filename 存在重复值，例如：{list(duplicates)!r}")

    return parsed_dates.dt.strftime("%Y%m%d")


def format_worksheet(writer: pd.ExcelWriter, sheet_name: str = "Sheet1") -> None:
    """给输出表添加简洁格式，不改变数据内容。"""
    worksheet = writer.book[sheet_name]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 28
    for column in range(2, worksheet.max_column + 1):
        worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = 14


def write_excel(df: pd.DataFrame, path: Path, overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise FileExistsError(f"输出已存在：{path}；如需覆盖请增加 --overwrite")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)
        format_worksheet(writer)


def assert_no_date_leakage(
    train_dates: set[str], val_dates: set[str], test_dates: set[str]
) -> None:
    if train_dates & val_dates:
        raise AssertionError("训练集与验证集存在日期重叠。")
    if train_dates & test_dates:
        raise AssertionError("训练集与测试集存在日期重叠。")
    if val_dates & test_dates:
        raise AssertionError("验证集与测试集存在日期重叠。")


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")

    output_dir = (
        args.output_dir.expanduser().resolve()
        if args.output_dir is not None
        else input_path.parent / "leave_one_date_out_2024"
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(input_path, sheet_name=parse_sheet(args.sheet))
    if df.empty:
        raise ValueError("输入工作表没有数据。")

    # 如果输入的是旧划分结果，丢弃旧 split 信息，重新按日期划分。
    data = df.drop(
        columns=[column for column in ("split", "data_split", "capture_date") if column in df],
        errors="ignore",
    ).copy()
    data["_capture_date"] = extract_dates(data, args.filename_col)

    unique_dates = sorted(data["_capture_date"].unique().tolist())
    if len(unique_dates) < 3:
        raise ValueError(
            "至少需要 3 个不同日期才能生成非空 train/val/test；"
            f"当前只有 {len(unique_dates)} 个日期：{unique_dates}"
        )

    source_filenames = set(data[args.filename_col].astype(str))
    fold_records: list[dict[str, object]] = []

    for fold_index, test_date in enumerate(unique_dates):
        # 下一个日期作为验证集；最后一个测试日期循环使用第一个日期验证。
        val_date = unique_dates[(fold_index + 1) % len(unique_dates)]
        train_dates = [
            date for date in unique_dates if date not in {test_date, val_date}
        ]

        train_mask = data["_capture_date"].isin(train_dates)
        val_mask = data["_capture_date"].eq(val_date)
        test_mask = data["_capture_date"].eq(test_date)

        split_frames = {
            "train": data.loc[train_mask].drop(columns="_capture_date").copy(),
            "val": data.loc[val_mask].drop(columns="_capture_date").copy(),
            "test": data.loc[test_mask].drop(columns="_capture_date").copy(),
        }

        split_date_sets = {
            split: set(frame[args.filename_col].astype(str).str[:8])
            for split, frame in split_frames.items()
        }
        assert_no_date_leakage(
            split_date_sets["train"],
            split_date_sets["val"],
            split_date_sets["test"],
        )
        if split_date_sets["test"] != {test_date}:
            raise AssertionError(f"第 {fold_index + 1} 折测试日期不正确。")

        combined_filenames = set().union(
            *(set(frame[args.filename_col].astype(str)) for frame in split_frames.values())
        )
        total_rows = sum(len(frame) for frame in split_frames.values())
        if total_rows != len(data) or combined_filenames != source_filenames:
            raise AssertionError(f"第 {fold_index + 1} 折没有完整覆盖源数据。")

        fold_name = f"fold_{fold_index + 1:02d}_test_{test_date}"
        fold_dir = output_dir / fold_name
        fold_dir.mkdir(parents=True, exist_ok=True)

        for split, frame in split_frames.items():
            write_excel(
                frame,
                fold_dir / SPLIT_FILENAMES[split],
                overwrite=args.overwrite,
            )

        fold_records.append(
            {
                "fold": fold_index + 1,
                "folder": fold_name,
                "test_date": test_date,
                "val_date": val_date,
                "train_dates": ",".join(train_dates),
                "train_samples": len(split_frames["train"]),
                "val_samples": len(split_frames["val"]),
                "test_samples": len(split_frames["test"]),
                "total_samples": total_rows,
            }
        )
        print(
            f"[{fold_index + 1:02d}/{len(unique_dates):02d}] "
            f"test={test_date}, val={val_date}, "
            f"train/val/test={len(split_frames['train'])}/"
            f"{len(split_frames['val'])}/{len(split_frames['test'])}"
        )

    summary = pd.DataFrame(fold_records)
    summary_path = output_dir / "leave_one_date_out_summary.xlsx"
    if summary_path.exists() and not args.overwrite:
        raise FileExistsError(
            f"输出已存在：{summary_path}；如需覆盖请增加 --overwrite"
        )

    date_counts = (
        data.groupby("_capture_date")
        .size()
        .rename("samples")
        .reset_index()
        .rename(columns={"_capture_date": "date"})
    )
    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="fold_summary", index=False)
        date_counts.to_excel(writer, sheet_name="date_counts", index=False)
        format_worksheet(writer, "fold_summary")
        format_worksheet(writer, "date_counts")
        writer.book["fold_summary"].column_dimensions["E"].width = 100

    config = {
        "input": str(input_path),
        "output_dir": str(output_dir),
        "method": "leave-one-date-out",
        "validation_rule": "next chronological date, cyclic",
        "total_samples": len(data),
        "dates": unique_dates,
        "folds": fold_records,
    }
    (output_dir / "split_config.json").write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"\n完成：共生成 {len(unique_dates)} 折，输出目录：{output_dir}")
    print("校验通过：每个日期测试一次，且每折 train/val/test 日期完全互斥。")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, FileExistsError, ValueError, AssertionError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        raise SystemExit(2) from exc
